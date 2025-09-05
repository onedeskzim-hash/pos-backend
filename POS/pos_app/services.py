import csv
import json
from io import StringIO, BytesIO
from datetime import datetime, timedelta
from django.core.mail import EmailMessage
from django.conf import settings
from django.template.loader import render_to_string
from django.utils import timezone
from .models import Transaction, Product, Customer, Reseller, Expense, ReportSchedule
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill


class ReportGenerator:
    @staticmethod
    def generate_business_report():
        """Generate comprehensive business report data"""
        transactions = Transaction.objects.all().order_by('-timestamp')
        products = Product.objects.all()
        customers = Customer.objects.all()
        resellers = Reseller.objects.all()
        expenses = Expense.objects.all()
        
        # Financial summary
        sold_transactions = transactions.filter(status='SOLD')
        total_revenue = sum(t.total_amount for t in sold_transactions)
        total_expenses = sum(e.amount for e in expenses)
        net_profit = total_revenue - total_expenses
        
        return {
            'transactions': transactions,
            'products': products,
            'customers': customers,
            'resellers': resellers,
            'expenses': expenses,
            'summary': {
                'total_revenue': total_revenue,
                'total_expenses': total_expenses,
                'net_profit': net_profit,
                'total_transactions': transactions.count(),
                'active_products': products.filter(is_active=True).count(),
                'active_customers': customers.filter(is_active=True).count(),
            }
        }
    
    @staticmethod
    def generate_excel_report():
        """Generate Excel report"""
        data = ReportGenerator.generate_business_report()
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Headers
        headers = ['Metric', 'Value']
        for col, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Summary data
        summary_data = [
            ['Total Revenue', f"${data['summary']['total_revenue']:.2f}"],
            ['Total Expenses', f"${data['summary']['total_expenses']:.2f}"],
            ['Net Profit', f"${data['summary']['net_profit']:.2f}"],
            ['Total Transactions', data['summary']['total_transactions']],
            ['Active Products', data['summary']['active_products']],
            ['Active Customers', data['summary']['active_customers']],
        ]
        
        for row, (metric, value) in enumerate(summary_data, 2):
            ws_summary.cell(row=row, column=1, value=metric)
            ws_summary.cell(row=row, column=2, value=value)
        
        # Transactions sheet
        ws_trans = wb.create_sheet("Transactions")
        trans_headers = ['Date', 'Receipt#', 'Product', 'Customer', 'Amount', 'Status', 'Payment Method']
        
        for col, header in enumerate(trans_headers, 1):
            cell = ws_trans.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        for row, transaction in enumerate(data['transactions'][:1000], 2):  # Limit to 1000 rows
            # Get customer/reseller name properly
            if transaction.customer:
                customer_name = transaction.customer.name
            elif transaction.reseller:
                customer_name = f"{transaction.reseller.name} (Reseller)"
            else:
                customer_name = 'Walk-in Customer'
            
            # Get all product names properly
            product_names = []
            if hasattr(transaction, 'items') and transaction.items.exists():
                items = list(transaction.items.all())
                product_names = [f"{item.product.name} (x{item.quantity})" for item in items]
            elif transaction.product:
                product_names = [f"{transaction.product.name} (x{transaction.quantity})"]
            
            product_name = "; ".join(product_names) if product_names else 'N/A'
            
            ws_trans.cell(row=row, column=1, value=transaction.timestamp.strftime('%Y-%m-%d'))
            ws_trans.cell(row=row, column=2, value=transaction.zimra_receipt_no or f'RCP-{transaction.id}')
            ws_trans.cell(row=row, column=3, value=product_name)
            ws_trans.cell(row=row, column=4, value=customer_name)
            ws_trans.cell(row=row, column=5, value=float(transaction.total_amount))
            ws_trans.cell(row=row, column=6, value=transaction.status)
            ws_trans.cell(row=row, column=7, value=transaction.payment_method)
        
        # Save to BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    @staticmethod
    def generate_pdf_report():
        """Generate PDF report (HTML for now)"""
        data = ReportGenerator.generate_business_report()
        
        html_content = f"""
        <html>
        <head>
            <title>Business Report - {datetime.now().strftime('%Y-%m-%d')}</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #f2f2f2; }}
                .summary {{ background: #f9f9f9; padding: 15px; margin: 20px 0; }}
            </style>
        </head>
        <body>
            <h1>Business Report - {datetime.now().strftime('%Y-%m-%d')}</h1>
            
            <div class="summary">
                <h2>Financial Summary</h2>
                <p>Total Revenue: ${data['summary']['total_revenue']:.2f}</p>
                <p>Total Expenses: ${data['summary']['total_expenses']:.2f}</p>
                <p>Net Profit: ${data['summary']['net_profit']:.2f}</p>
                <p>Total Transactions: {data['summary']['total_transactions']}</p>
            </div>
            
            <h2>Recent Transactions</h2>
            <table>
                <tr>
                    <th>Date</th><th>Product</th><th>Customer</th><th>Amount</th><th>Status</th>
                </tr>
        """
        
        for transaction in data['transactions'][:50]:  # Limit to 50 rows
            # Get customer/reseller name properly
            if transaction.customer:
                customer_name = transaction.customer.name
            elif transaction.reseller:
                customer_name = f"{transaction.reseller.name} (Reseller)"
            else:
                customer_name = 'Walk-in Customer'
            
            # Get all product names properly
            product_names = []
            if hasattr(transaction, 'items') and transaction.items.exists():
                items = list(transaction.items.all())
                product_names = [f"{item.product.name} (x{item.quantity})" for item in items]
            elif transaction.product:
                product_names = [f"{transaction.product.name} (x{transaction.quantity})"]
            
            product_name = "; ".join(product_names) if product_names else 'N/A'
            
            html_content += f"""
                <tr>
                    <td>{transaction.timestamp.strftime('%Y-%m-%d')}</td>
                    <td>{product_name}</td>
                    <td>{customer_name}</td>
                    <td>${transaction.total_amount:.2f}</td>
                    <td>{transaction.status}</td>
                </tr>
            """
        
        html_content += """
            </table>
        </body>
        </html>
        """
        
        return html_content.encode('utf-8')


class EmailService:
    @staticmethod
    def send_scheduled_report(schedule_id):
        """Send scheduled report via email"""
        try:
            schedule = ReportSchedule.objects.get(id=schedule_id, is_active=True)
            
            # Generate report
            if schedule.format == 'excel':
                report_data = ReportGenerator.generate_excel_report()
                filename = f'business_report_{datetime.now().strftime("%Y%m%d")}.xlsx'
                content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            elif schedule.format == 'pdf':
                report_data = ReportGenerator.generate_pdf_report()
                filename = f'business_report_{datetime.now().strftime("%Y%m%d")}.html'
                content_type = 'text/html'
            else:  # json
                data = ReportGenerator.generate_business_report()
                # Convert to JSON-serializable format
                json_data = {
                    'summary': data['summary'],
                    'transactions_count': data['transactions'].count(),
                    'generated_at': datetime.now().isoformat()
                }
                report_data = json.dumps(json_data, indent=2).encode('utf-8')
                filename = f'business_report_{datetime.now().strftime("%Y%m%d")}.json'
                content_type = 'application/json'
            
            # Create email
            subject = f'Scheduled Business Report - {datetime.now().strftime("%Y-%m-%d")}'
            message = f"""
            Dear User,
            
            Please find attached your scheduled {schedule.frequency} business report.
            
            Report Details:
            - Format: {schedule.format.upper()}
            - Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}
            - Frequency: {schedule.frequency.title()}
            
            Best regards,
            GiveSolar-POS
            """
            
            email = EmailMessage(
                subject=subject,
                body=message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=[schedule.email]
            )
            
            email.attach(filename, report_data, content_type)
            email.send()
            
            # Update schedule
            schedule.last_sent = timezone.now()
            schedule.save()
            
            return True
            
        except Exception as e:
            print(f"Error sending scheduled report: {e}")
            return False